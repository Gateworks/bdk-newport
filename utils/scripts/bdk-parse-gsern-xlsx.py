#!/usr/bin/env python2
# ***********************license start***********************************
# Copyright (C) 2018 Marvell International Ltd.
# SPDX-License-Identifier: BSD-3-Clause
# https://spdx.org/licenses
# **********************license end**************************************

import os
import sys
import openpyxl
import datetime
import pprint
import types

pp = pprint.PrettyPrinter(indent=4, width=132)
chip_db = None

def load_db():
    inf = open(os.environ["ASIM"] + "/tools/csr-tool/db/csr-cn93xx.db")
    db = eval(inf.read())
    inf.close()
    return db

#
# Load the XLSX files from a directory and return them. Some
# very simple sanity checks are done on them to detect format
# changes, but no parsing yet
#
def load_xlsx(basedir):
    wb_tx_eq = openpyxl.load_workbook(filename = basedir + "/tx_equalizer_defaults_by_bitrate.xlsx", read_only=True, data_only=True)
    # Sanity checks to make sure format didn't change
    assert "Sheet1" in wb_tx_eq.sheetnames, wb_tx_eq.sheetnames
    ws = wb_tx_eq["Sheet1"]
    assert ws["D4"].value == "Data Rate (Gbd)", ws["D4"].value
    assert ws["D6"].value == "Standard", ws["D6"].value
    assert ws["C19"].value == "Values in hex", ws["C19"].value
    assert ws["D20"].value == "div20", ws["D20"].value

    wb_rx_eq = openpyxl.load_workbook(filename = basedir + "/rx_equalizer_defaults_by_bitrate.xlsx", read_only=True, data_only=True)
    assert "Sheet1" in wb_rx_eq.sheetnames, wb_rx_eq.sheetnames
    ws = wb_rx_eq["Sheet1"]
    assert ws["D4"].value == "Data Rate (Gbd)", ws["D4"].value
    assert ws["D6"].value == "Standard", ws["D6"].value
    assert ws["C19"].value == "Values in hex", ws["C19"].value
    assert ws["D20"].value == "c1_q_adjust", ws["D20"].value

    wb_rx_eq_lab = openpyxl.load_workbook(filename = basedir + "/rx_equalizer_defaults_by_bitrate_lab.xlsx", read_only=True, data_only=True)
    assert "Sheet1" in wb_rx_eq_lab.sheetnames, wb_rx_eq_lab.sheetnames
    ws = wb_rx_eq_lab["Sheet1"]
    assert ws["D4"].value == "Data Rate (Gbd)", ws["D4"].value
    assert ws["D6"].value == "Standard", ws["D6"].value
    assert ws["C19"].value == "Values in hex", ws["C19"].value
    assert ws["D20"].value == "c1_q_adjust", ws["D20"].value

    wb_idle = openpyxl.load_workbook(filename = basedir + "/idle_refset_by_protocols.xlsx", read_only=True, data_only=True)
    assert "Ethernet" in wb_idle.sheetnames, wb_idle.sheetnames
    assert "Supported" in wb_idle.sheetnames, wb_idle.sheetnames
    assert "RTL Lean Version" in wb_idle.sheetnames, wb_idle.sheetnames
    ws = wb_idle["Ethernet"]
    return wb_tx_eq, wb_rx_eq, wb_rx_eq_lab, wb_idle

def convert_to_python(ws):
    xlsx_data = []
    for row in ws:
        xlsx_data.append([])
        for col in row:
            data = col.value
            try:
                data = data.encode("ascii", "replace")
            except:
                pass
            xlsx_data[-1].append(data)
    return xlsx_data

#
# Parse the transmit settings
#
def parse_xlsx_tx_eq(wb_tx_eq, gsern_data):
    xlsx_data = convert_to_python(wb_tx_eq["Sheet1"])
    serdes_modes = []
    assert xlsx_data[3][3] == "Data Rate (Gbd)"
    assert xlsx_data[5][3] == "Standard"
    assert xlsx_data[1][30] == "LAB USE ONLY"
    assert xlsx_data[1][50] == "LAB USE ONLY"
    assert xlsx_data[1][67] == "LAB USE ONLY"
    assert xlsx_data[1][74] == "OCI"
    ws_row_rate = xlsx_data[3]
    ws_row_standard = xlsx_data[5]
    # Mark LAB only speeds
    for c in range(30,74):
        ws_row_standard[c] = "LAB"
    ws_row_standard[49] = None
    ws_row_standard[66] = None
    for c in range(74,len(ws_row_standard)):
        ws_row_standard[c] = "OCI"
    column = 4
    while column < len(ws_row_rate):
        rate_float = ws_row_rate[column]
        if not rate_float:
            break
        rate_hz = int(rate_float * 1000000000)
        assert(rate_hz == rate_float * 1000000000)
        standard = ws_row_standard[column]
        if standard:
            if standard.startswith("LAUI"):
                standard = "GEN"
            else:
                standard = standard.upper()
        elif rate_hz == 1250000000:
            standard = "SGMII"
        elif rate_hz == 1228000000:
            standard = "CIPRI"
        elif rate_hz in [2500000000, 5000000000, 8000000000, 16000000000]:
            standard = "PCIE"
        elif rate_hz in [1500000000, 3000000000, 6000000000]:
            standard = "SATA"
        else:
            standard = "GEN"
        mode = "%s_%011d" % (standard, rate_hz)
        serdes_modes.append(mode)
        column = column + 1

    assert xlsx_data[18][2] == "Values in hex"
    row = 19
    while row < len(xlsx_data):
        field = xlsx_data[row][3]
        if not field:
            break
        field = field.replace(" (", "(")
        field = field.replace(" ", "_")
        for col,mode in enumerate(serdes_modes):
            if mode.startswith("LAB") or mode.startswith("OCI"):
                continue
            column = 4 + col
            value = str(xlsx_data[row][column])
            value = int(value, 16)
            if not field in gsern_data:
                gsern_data[field] = {}
            assert not mode in gsern_data[field], "%s %s" % (mode, gsern_data[field])
            gsern_data[field][mode] = value
        row = row + 1
    return 0

#
# Parse the receive settings
#
def parse_xlsx_rx_eq(wb_rx_eq, gsern_data):
    # RX and TX are the same format, so use TX parser
    return parse_xlsx_tx_eq(wb_rx_eq, gsern_data)

#
# Parse the idle settings
#
def parse_xlsx_idle(wb_idle, gsern_data):
    return 0

#
# Parse the XLSX file in a dir
#
def parse_xlsx(basedir):
    wb_tx_eq, wb_rx_eq, wb_rx_eq_lab, wb_idle = load_xlsx(basedir)
    gsern_data = {}
    parse_xlsx_tx_eq(wb_tx_eq, gsern_data)
    #parse_xlsx_rx_eq(wb_rx_eq, gsern_data)
    parse_xlsx_rx_eq(wb_rx_eq_lab, gsern_data)
    parse_xlsx_idle(wb_idle, gsern_data)
    return gsern_data

#
# Lookup spreadsheet names and find the correct CSR name and field
#
def lookup_gsern_csrs(gsern_data):
    global chip_db
    csr_map = {}
    fields = gsern_data.keys()
    fields.sort()
    for field in fields:
        found = False
        last_found = None
        csr_names = chip_db.keys()
        csr_names.sort()
        for csr_name in csr_names:
            if not csr_name.startswith("GSERN#_LANE#_") and not csr_name.startswith("GSERN#_COMMON_"):
                continue
            csr = chip_db[csr_name]
            if not "fields" in csr:
                continue
            for f in csr["fields"]:
                no_phase = field.split("(")[0]
                if no_phase == f["name"]:
                    if found:
                        if last_found == "GSERN#_LANE#_RST2_BCFG":
                            continue
                        if last_found == "GSERN#_LANE#_TX_1_BCFG":
                            continue
                        if not csr_name in ["GSERN#_LANE#_RST2_BCFG", "GSERN#_LANE#_TX_1_BCFG"]:
                            print csr_name, field, last_found
                            assert False
                    found = True
                    last_found = csr_name
                    csr_map[field] = "%s[%s]" % (csr_name, f["name"])
        assert found, field
    return csr_map

def combine_chips(t93, t95):
    both = {}
    keys = t93.keys()
    keys.sort()
    for k in keys:
        both[k] = {}
        for speed in t93[k]:
            if t93[k][speed] == t95[k][speed]:
                both[k][speed] = t93[k][speed]
            else:
                both[k][speed] = {}
                both[k][speed]["cn96xx"] = t93[k][speed]
                both[k][speed]["cnf95xx"] = t95[k][speed]
                #print k, speed, both[k][speed]
    return both
#
# Write the Copyright banner
#
def writeCopyrightBanner(out):
    year = datetime.date.today().year
    out.write("/***********************license start***********************************\n")
    out.write("* Copyright (C) %d Marvell International Ltd.\n" % year)
    out.write("* SPDX-License-Identifier: BSD-3-Clause\n")
    out.write("* https://spdx.org/licenses\n")
    out.write("***********************license end**************************************/\n")
    out.write("\n")

#
# Write header for GSERN settings tables
#
def write_tables(gsern_data, csr_map):
    outf = open("output/octeontx/cavm-gsern-settings.h", "w")
    outf.write("#ifndef _CAVM_GSERN_SETTINGS_H_\n")
    outf.write("#define _CAVM_GSERN_SETTINGS_H_\n")
    outf.write("/* This file is auto-generated. Do not edit */\n")
    outf.write("\n")
    writeCopyrightBanner(outf)
    fields = gsern_data.keys()
    fields.sort()
    modes = gsern_data[fields[0]].keys()
    modes.sort()
    outf.write("/* Enumaration used to index arrays below */\n")
    outf.write("enum gsern_lane_modes {\n")
    for mode in modes:
        if mode == modes[-1]:
            outf.write("    GSERN_%s\n" % mode)
        else:
            outf.write("    GSERN_%s,\n" % mode)
    outf.write("};\n")
    outf.write("\n")
    outf.write("/* GSERN field settings indexed by enum gsern_data_modes */\n")
    outf.write("\n")
    for field in fields:
        all_same = True
        base_mode = "GEN_02457600000"
        for mode in modes:
            if not mode in gsern_data[field]:
                continue
            a = str(gsern_data[field][mode])
            b = str(gsern_data[field][base_mode])
            all_same = all_same and (a == b)
        c_name = field.replace("(", "_")
        c_name = c_name.replace(")", "")
        if all_same:
            if isinstance(gsern_data[field][base_mode], types.DictType):
                d = gsern_data[field][base_mode]
                outf.write("static const int gsern_lane_%s_%s = 0x%x; /* %s, same for all modes */\n" % (c_name, "cn96xx", d["cn96xx"], csr_map[field]))
                outf.write("static const int gsern_lane_%s_%s = 0x%x; /* %s, same for all modes */\n" % (c_name, "cnf95xx", d["cnf95xx"], csr_map[field]))
            else:
                outf.write("static const int gsern_lane_%s = 0x%x; /* %s, same for all modes */\n" % (c_name, gsern_data[field][base_mode], csr_map[field]))
        else:
            per_chip = False
            for mode in modes:
                if isinstance(gsern_data[field][mode], types.DictType):
                    per_chip = True
                    break
            if per_chip:
                for chip in ["cn96xx", "cnf95xx"]:
                    outf.write("static const int gsern_lane_%s_%s[] = {" % (c_name, chip))
                    for mode in modes:
                        if mode in gsern_data[field]:
                            if isinstance(gsern_data[field][mode], types.DictType):
                                v = gsern_data[field][mode][chip]
                            else:
                                v = gsern_data[field][mode]
                        else:
                            assert False, "Missing %s[%s]" % (field, mode)
                        if mode == modes[-1]:
                            outf.write(" 0x%x" % v)
                        else:
                            outf.write(" 0x%x," % v)
                    outf.write("}; /* %s */\n" % csr_map[field])
            else:
                outf.write("static const int gsern_lane_%s[] = {" % c_name)
                for mode in modes:
                    if mode in gsern_data[field]:
                        v = gsern_data[field][mode]
                    else:
                        assert False, "Missing %s[%s]" % (field, mode)
                    if mode == modes[-1]:
                        outf.write(" 0x%x" % v)
                    else:
                        outf.write(" 0x%x," % v)
                outf.write("}; /* %s */\n" % csr_map[field])
    outf.write("\n")
    outf.write("/* End of generated data */\n")
    outf.write("#endif /* _CAVM_GSERN_SETTINGS_H_ */\n")
    outf.close()
#
# Main entry point
#
def main():
    global chip_db

    # Load the CSR database
    chip_db = load_db()
    gsern_data_t93 = parse_xlsx("t93_config_spreadsheets")
    gsern_data_t95 = parse_xlsx("t95_config_spreadsheets")
    csr_map = lookup_gsern_csrs(gsern_data_t93)
    both = combine_chips(gsern_data_t93, gsern_data_t95)
    write_tables(both, csr_map)
    return 0

sys.exit(main())
